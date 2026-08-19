#!/usr/bin/env python3
"""Download the constituents of an ETF from the official issuer documents and export
them to a normalized XLSX.

Supported issuers: iShares (BlackRock), Xtrackers (DWS), Vanguard, SPDR (State Street).

    python etf_constituents.py IE00B4L5Y983
    python etf_constituents.py LU0397221945 -o portfolio.xlsx --enrich-ticker

Columns produced:
    Ticker, ISIN, Name, Sector, Class, Country, Region, Category, Currency, Weight

Weights are expressed in percentage points and always add up to 100: any residual
(issuer rounding, positions not broken down) ends up in an "Other" balancing row.
"""

from __future__ import annotations

import argparse
import datetime as dt
import logging
import os
import re
import sys
import time

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

import currency
import msci
import taxonomy
from providers import Fund, Holding, ProviderError, ProviderRegistry, is_valid_isin

log = logging.getLogger("etf_constituents")

COLUMNS = [
    "Ticker",
    "ISIN",
    "Name",
    "Sector",
    "Class",
    "Country",
    "Region",
    "Category",
    "Currency",
    "Weight",
]

BALANCING_LABEL = "Weight balance"
# Set on rows produced by expanding a nested ETF, and rendered into Name at the end.
# Deliberately outside COLUMNS: write_xlsx and aggregate both iterate COLUMNS, so the
# extra key rides along without being written or overwritten.
SOURCE_KEY = "_Source"
WEIGHT_TOLERANCE = 1e-9
# Below this threshold the residual is issuer rounding noise and the balancing row is
# added silently; above it, it is worth reporting.
MATERIAL_RESIDUAL = 0.1

# An ISIN is a candidate for expansion only if it passes this filter: without it, an
# equity fund would fire ~1300 HTTP probes to find out they are all single stocks.
_FUND_NAME_RE = re.compile(
    r"UCITS ETF|\biShares\b|\bXtrackers\b|\bSPDR\b|\bState Street\b|\bETF\b|\bINDEX FUND\b",
    re.I,
)
_FUND_DOMICILES = ("IE", "LU")


class Row(dict):
    """An already normalized output row."""


def build_row(holding: Holding) -> Row:
    country, region, category = msci.classify(holding.country_raw)
    return Row(
        Ticker=holding.ticker,
        ISIN=holding.isin,
        Name=holding.name,
        Sector=taxonomy.normalize_sector(holding.sector_raw),
        Class=taxonomy.normalize_class(holding.asset_class_raw),
        Country=country,
        Region=region,
        Category=category,
        # Not every issuer prices its holdings: fall back to the country's currency.
        Currency=holding.currency or currency.from_country(holding.country_raw),
        Weight=holding.weight,
    )


# ---------------------------------------------------------------------------
# Recursive expansion of nested ETFs
# ---------------------------------------------------------------------------


def looks_like_fund(row: Row) -> bool:
    if not is_valid_isin(row["ISIN"]):
        return False
    if row["Class"] == taxonomy.FUND:
        return True
    if _FUND_NAME_RE.search(row["Name"] or ""):
        return True
    # Last case: an IE/LU domiciled vehicle the issuer was not able to classify.
    return row["ISIN"][:2] in _FUND_DOMICILES and row["Sector"] == taxonomy.OTHER


def collect_rows(
    fund: Fund,
    registry: ProviderRegistry,
    *,
    expand: bool,
    max_depth: int,
    depth: int = 0,
    visited: frozenset[str] = frozenset(),
) -> list[Row]:
    """Leaf rows of the fund, with the sub-fund weights already rescaled."""
    rows: list[Row] = []
    for holding in fund.holdings:
        row = build_row(holding)

        # A negative max_depth means unlimited: cycles are already bounded by `visited`
        # and by the provider miss cache, so the depth cap is not what makes this end.
        if not expand or (0 <= max_depth <= depth):
            rows.append(row)
            continue
        if row["ISIN"] in visited or row["ISIN"] == fund.isin.upper():
            rows.append(row)  # cycle: keep the parent, the weight is not lost
            continue
        if not looks_like_fund(row):
            rows.append(row)
            continue

        try:
            child = registry.fetch(row["ISIN"])
        except ProviderError:
            rows.append(row)
            continue

        child_name = child.name or row["Name"]
        log.info(
            "  expanding %s %s (%.4f%%) -> %d constituents",
            row["ISIN"],
            child_name,
            row["Weight"],
            len(child.holdings),
        )
        child_rows = collect_rows(
            child,
            registry,
            expand=expand,
            max_depth=max_depth,
            depth=depth + 1,
            visited=visited | {row["ISIN"]},
        )
        child_total = sum(r["Weight"] for r in child_rows)
        if child_total <= 0:
            rows.append(row)
            continue

        # Rescale on the actual child total: if its weights do not add up to 100,
        # normalizing them here avoids losing or inflating the parent weight.
        factor = row["Weight"] / child_total
        for child_row in child_rows:
            scaled = dict(child_row)
            scaled["Weight"] = child_row["Weight"] * factor
            # setdefault, not assignment: the deepest level got there first, and the row
            # should name the ETF it came from directly, not every one above it.
            scaled.setdefault(SOURCE_KEY, child_name)
            rows.append(Row(scaled))

    return rows


# ---------------------------------------------------------------------------
# Aggregation and balancing
# ---------------------------------------------------------------------------


def aggregate(rows: list[Row]) -> list[Row]:
    """Sum the weights of the rows describing the same security in the same source ETF.

    The source is part of the key because a security reached through several sub-funds
    has to keep them apart: merging on ISIN alone would leave one arbitrary ETF name on a
    weight that came from all of them.
    """
    merged: dict[tuple, Row] = {}
    for row in rows:
        source = row.get(SOURCE_KEY, "")
        key = (row["ISIN"], source) if row["ISIN"] else ("", row["Name"], row["Currency"], source)
        existing = merged.get(key)
        if existing is None:
            merged[key] = Row(row)
            continue
        existing["Weight"] += row["Weight"]
        for column in COLUMNS:
            if column != "Weight" and not existing[column] and row[column]:
                existing[column] = row[column]
    return list(merged.values())


def label_sources(rows: list[Row]) -> list[Row]:
    """Append the source ETF, in brackets, to the Name of every expanded row."""
    for row in rows:
        source = row.get(SOURCE_KEY)
        if source:
            row["Name"] = f"{row['Name']} ({source})".strip()
    return rows


def add_balancing_row(rows: list[Row]) -> list[Row]:
    """Add the `Other` row that brings the total to exactly 100%."""
    total = sum(row["Weight"] for row in rows)
    residual = 100.0 - total
    if abs(residual) <= WEIGHT_TOLERANCE:
        return rows
    if abs(residual) > MATERIAL_RESIDUAL:
        log.warning(
            "the issuer weights add up to %.4f%%: the balancing row is worth %+.4f%%",
            total,
            residual,
        )
    rows.append(
        Row(
            Ticker="",
            ISIN="",
            Name=BALANCING_LABEL,
            Sector=taxonomy.OTHER,
            Class=taxonomy.OTHER,
            Country=msci.OTHER,
            Region=msci.OTHER,
            Category=msci.OTHER,
            Currency="",
            Weight=residual,
        )
    )
    return rows


def sort_rows(rows: list[Row]) -> list[Row]:
    """Descending weight, with the balancing row always last."""
    return sorted(rows, key=lambda r: (r["Name"] == BALANCING_LABEL and not r["ISIN"], -r["Weight"]))


# ---------------------------------------------------------------------------
# Ticker enrichment (OpenFIGI, opt-in)
# ---------------------------------------------------------------------------

OPENFIGI_URL = "https://api.openfigi.com/v3/mapping"


def enrich_tickers(rows: list[Row], session: requests.Session) -> int:
    """Resolve the missing Tickers via OpenFIGI. Returns the number of tickers added."""
    pending = [r for r in rows if not r["Ticker"] and is_valid_isin(r["ISIN"])]
    if not pending:
        return 0

    api_key = os.environ.get("OPENFIGI_API_KEY")
    headers = {"Content-Type": "application/json"}
    if api_key:
        headers["X-OPENFIGI-APIKEY"] = api_key
    # Public limits: 10 jobs/request and 25 requests/minute without an API key,
    # 100 jobs/request and 25 requests/6s with an API key.
    batch_size = 100 if api_key else 10
    min_interval = 6.0 / 25 if api_key else 60.0 / 25

    by_isin: dict[str, list[Row]] = {}
    for row in pending:
        by_isin.setdefault(row["ISIN"], []).append(row)
    isins = list(by_isin)

    log.info(
        "OpenFIGI: resolving %d ISINs in %d requests (~%.0fs)",
        len(isins),
        -(-len(isins) // batch_size),
        -(-len(isins) // batch_size) * min_interval,
    )

    resolved = 0
    for start in range(0, len(isins), batch_size):
        batch = isins[start : start + batch_size]
        try:
            response = session.post(
                OPENFIGI_URL,
                json=[{"idType": "ID_ISIN", "idValue": isin} for isin in batch],
                headers=headers,
                timeout=(10, 60),
            )
            if response.status_code == 429:
                log.warning("OpenFIGI rate limit reached: stopping the enrichment")
                break
            response.raise_for_status()
            payload = response.json()
        except (requests.RequestException, ValueError) as exc:
            log.warning("OpenFIGI: request failed (%s), carrying on without it", exc)
            break

        for isin, result in zip(batch, payload):
            data = (result or {}).get("data") or []
            if not data:
                continue
            ticker = (data[0].get("ticker") or "").strip()
            if ticker:
                for row in by_isin[isin]:
                    row["Ticker"] = ticker
                resolved += 1

        if start + batch_size < len(isins):
            time.sleep(min_interval)

    log.info("OpenFIGI: %d/%d tickers resolved", resolved, len(isins))
    return resolved


# ---------------------------------------------------------------------------
# Output XLSX
# ---------------------------------------------------------------------------


def write_xlsx(path: str, rows: list[Row]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Constituents"

    sheet.append(COLUMNS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"

    weight_column = COLUMNS.index("Weight") + 1
    for row in rows:
        sheet.append([row[column] for column in COLUMNS])
        # Excel treats the percentage format as a fraction: 0.0558 rendered as "5.58%".
        cell = sheet.cell(row=sheet.max_row, column=weight_column)
        cell.value = row["Weight"] / 100.0
        cell.number_format = "0.0000%"

    widths = {"Ticker": 12, "ISIN": 15, "Name": 42, "Sector": 24, "Class": 14,
              "Country": 22, "Region": 24, "Category": 18, "Currency": 10, "Weight": 12}
    for index, column in enumerate(COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = widths[column]
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{sheet.max_row}"

    workbook.save(path)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Download the constituents of an ETF "
            "(iShares, Xtrackers, Vanguard, SPDR) to XLSX."
        ),
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("isin", help="ISIN of the ETF")
    parser.add_argument("-o", "--output", help="destination XLSX file")
    parser.add_argument(
        "--no-expand",
        action="store_true",
        help="do not expand constituents that are themselves ETFs",
    )
    parser.add_argument(
        "--max-depth",
        type=int,
        default=-1,
        help="maximum expansion depth, negative for unlimited",
    )
    parser.add_argument(
        "--enrich-ticker",
        action="store_true",
        help="resolve the missing Tickers via OpenFIGI (slow, third-party call)",
    )
    parser.add_argument("-v", "--verbose", action="store_true", help="detailed logging")
    return parser.parse_args(argv)


def default_output_name(isin: str, as_of: str) -> str:
    stamp = re.sub(r"[^A-Za-z0-9]+", "-", as_of).strip("-") if as_of else ""
    stamp = stamp or dt.date.today().isoformat()
    return f"{isin}_constituents_{stamp}.xlsx"


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(message)s",
        stream=sys.stderr,
    )

    isin = args.isin.strip().upper()
    if not is_valid_isin(isin):
        log.error("'%s' is not a valid ISIN", args.isin)
        return 2

    registry = ProviderRegistry()
    try:
        fund = registry.fetch(isin)
    except ProviderError as exc:
        log.error("%s", exc)
        return 1

    fund_name = registry.resolve_name(fund)
    log.info("%s - %s (%s), %d constituents", isin, fund_name or "?", fund.issuer, len(fund.holdings))

    rows = collect_rows(
        fund,
        registry,
        expand=not args.no_expand,
        max_depth=args.max_depth,
    )
    rows = aggregate(rows)
    rows = label_sources(rows)
    rows = add_balancing_row(rows)
    rows = sort_rows(rows)

    if args.enrich_ticker:
        enrich_tickers(rows, registry.session)

    taxonomy.report_unknown()
    msci.report_unresolved()
    currency.report_unmapped()

    output = args.output or default_output_name(isin, fund.as_of)
    write_xlsx(output, rows)

    total = sum(row["Weight"] for row in rows)
    log.info("Wrote %d rows to %s (weights total %.6f%%)", len(rows), output, total)
    return 0


if __name__ == "__main__":
    sys.exit(main())
