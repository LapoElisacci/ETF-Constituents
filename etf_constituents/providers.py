"""Download of constituents from the official documents of the supported issuers.

Issuers: iShares (BlackRock), Xtrackers (DWS), Vanguard, SPDR (State Street) and Amundi,
plus UBS from a workbook the user downloads (see UbsFileProvider).

Every provider exposes `fetch(isin)` and returns a `Fund` whose weights are already
expressed in percentage points, leaving sector/asset class/country in the raw issuer
form: normalization happens downstream in `taxonomy` and `msci`.
"""

from __future__ import annotations

import csv
import io
import logging
import re
import warnings
from dataclasses import dataclass, field

import openpyxl
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

import taxonomy

log = logging.getLogger(__name__)

USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/126.0 Safari/537.36"
)
TIMEOUT = (10, 120)  # (connect, read): full exports can be slow

ISIN_RE = re.compile(r"^[A-Z]{2}[A-Z0-9]{9}[0-9]$")


class ProviderError(RuntimeError):
    """Recoverable error while downloading/parsing the constituents of a fund."""


@dataclass
class Holding:
    """A constituent, with fields still in the issuer nomenclature."""

    isin: str = ""
    ticker: str = ""
    name: str = ""
    sector_raw: str = ""
    asset_class_raw: str = ""
    country_raw: str = ""
    currency: str = ""
    weight: float = 0.0  # percentage points (5.58 = 5.58%)
    exchange: str = ""


@dataclass
class Fund:
    isin: str
    name: str
    issuer: str
    as_of: str = ""
    holdings: list[Holding] = field(default_factory=list)


def build_session() -> requests.Session:
    session = requests.Session()
    session.headers.update({"User-Agent": USER_AGENT, "Accept-Encoding": "gzip, deflate"})
    # 500 is not in the forcelist: DWS uses it for "unknown ISIN", retrying it
    # does not help and slows down issuer detection.
    retry = Retry(
        total=3,
        backoff_factor=0.5,
        status_forcelist=(429, 502, 503, 504),
        allowed_methods=frozenset({"GET", "POST"}),
    )
    adapter = HTTPAdapter(max_retries=retry, pool_maxsize=16)
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    return session


def is_valid_isin(value: str | None) -> bool:
    return bool(value and ISIN_RE.match(value.strip().upper()))


# ---------------------------------------------------------------------------
# iShares / BlackRock
# ---------------------------------------------------------------------------


class IsharesProvider:
    name = "iShares"

    SEARCH_URL = "https://www.ishares.com/varnish-api/core-search/search/products"
    DATA_URL = (
        "https://www.ishares.com/varnish-api/uk-retail01-product-data"
        "/product-data/api/v2/get-product-data"
    )
    # (targetSite, locale): the first one exposing a fund with the requested ISIN wins.
    SITES = (("ishares-uk", "en_GB"), ("us-ishares", "en_US"))
    MAX_CANDIDATES = 5

    def __init__(self, session: requests.Session):
        self.session = session

    def _product_data(self, portfolio_id: str, site: str, locale: str, component: str) -> dict:
        params = {
            "appType": "PRODUCT_PAGE",
            "appSubType": "ISHARES",
            "targetSite": site,
            "locale": locale,
            "userType": "individual",
            "excludeContent": "true",
            "includeConfig": "true",
            "component": component,
            "portfolioId": portfolio_id,
        }
        response = self.session.get(
            self.DATA_URL,
            params=params,
            headers={"Accept": "application/json", "x-application-id": "pp-ui-csr"},
            timeout=TIMEOUT,
        )
        response.raise_for_status()
        return response.json()

    def _search(self, isin: str, site: str, locale: str) -> list[dict]:
        response = self.session.get(
            self.SEARCH_URL,
            params={"site": site, "locale": locale, "query": isin},
            headers={"Accept": "application/json"},
            timeout=TIMEOUT,
        )
        if response.status_code != 200:
            return []
        return response.json().get("results") or []

    def resolve(self, isin: str) -> tuple[str, str, str, str] | None:
        """ISIN -> (portfolioId, fundName, targetSite, locale), or None.

        The search is fuzzy (it returns results even for an unknown ISIN), so every
        candidate must be confirmed by reading the ISIN from `keyFundFacts`.
        """
        isin = isin.strip().upper()
        for site, locale in self.SITES:
            try:
                results = self._search(isin, site, locale)
            except requests.RequestException as exc:
                log.debug("iShares search failed on %s: %s", site, exc)
                continue

            seen: set[str] = set()
            candidates: list[dict] = []
            for item in results:
                pid = str(item.get("portfolioId") or "")
                if pid and pid not in seen:
                    seen.add(pid)
                    candidates.append(item)

            for item in candidates[: self.MAX_CANDIDATES]:
                pid = str(item["portfolioId"])
                try:
                    facts = self._product_data(pid, site, locale, "keyFundFacts")
                except (requests.RequestException, ValueError) as exc:
                    log.debug("keyFundFacts failed for %s: %s", pid, exc)
                    continue
                if _key_fund_fact_isin(facts) == isin:
                    fund_name = facts.get("fundName") or item.get("fundName") or isin
                    return pid, fund_name, site, locale
        return None

    def fetch(self, isin: str) -> Fund:
        resolved = self.resolve(isin)
        if resolved is None:
            raise ProviderError(f"{isin} does not appear to be an iShares ETF")
        portfolio_id, fund_name, site, locale = resolved

        try:
            payload = self._product_data(portfolio_id, site, locale, "holdings")
        except (requests.RequestException, ValueError) as exc:
            raise ProviderError(f"iShares constituents download failed for {isin}: {exc}") from exc

        try:
            points = payload["componentsByNameMap"]["holdings"]["containersByNameMap"]["all"][
                "dataPointsByNameMap"
            ]
        except (KeyError, TypeError) as exc:
            raise ProviderError(f"unexpected holdings structure for {isin}") from exc

        def column(field_name: str, *, numeric: bool = False) -> list:
            entry = points.get(field_name) or {}
            # `formattedValue` is rounded to 2 decimals: on funds with thousands of
            # positions the accumulated error is worth whole percentage points (on IEAC
            # 101.79 against 99.9998). For numbers, `value` is the field to read.
            keys = ("value", "formattedValue") if numeric else ("formattedValue", "value")
            for key in keys:
                values = entry.get(key)
                if isinstance(values, list):
                    return values
            return []

        isins = column("isin")
        if not isins:
            raise ProviderError(f"no constituent returned for {isin}")

        tickers = column("ticker")
        names = column("issueName")
        sectors = column("sectorName")
        classes = column("assetClass")
        countries = column("countryOfRisk")
        currencies = column("marketCurrencyCode")
        weights = column("holdingPercent", numeric=True)
        exchanges = column("exchange")

        def at(values: list, index: int) -> str:
            if index < len(values):
                value = values[index]
                return "" if value is None else str(value).strip()
            return ""

        holdings = []
        for i in range(len(isins)):
            holdings.append(
                Holding(
                    isin=at(isins, i).upper(),
                    ticker=at(tickers, i),
                    name=at(names, i),
                    sector_raw=at(sectors, i),
                    asset_class_raw=at(classes, i),
                    country_raw=at(countries, i),
                    currency=at(currencies, i),
                    weight=_to_float(at(weights, i)),
                    exchange=at(exchanges, i),
                )
            )

        as_of = ""
        as_of_entry = points.get("asOfDate") or {}
        if isinstance(as_of_entry.get("formattedValue"), str):
            as_of = as_of_entry["formattedValue"]

        return Fund(isin=isin, name=fund_name, issuer=self.name, as_of=as_of, holdings=holdings)


def _key_fund_fact_isin(payload: dict) -> str:
    try:
        containers = payload["componentsByNameMap"]["keyFundFacts"]["containersByNameMap"]
    except (KeyError, TypeError):
        return ""
    for container in containers.values():
        entry = (container.get("dataPointsByNameMap") or {}).get("isin") or {}
        value = entry.get("value") or entry.get("formattedValue")
        if isinstance(value, str) and value.strip():
            return value.strip().upper()
    return ""


# ---------------------------------------------------------------------------
# Xtrackers / DWS
# ---------------------------------------------------------------------------


class XtrackersProvider:
    name = "Xtrackers"

    EXPORT_URL = "https://etf.dws.com/etfdata/export/GBR/ENG/csv/product/constituent/{isin}/"
    PRODUCT_URL = "https://etf.dws.com/en-gb/{isin}/"
    # Acronyms that must not go through .title() when rebuilding the name from the slug.
    _ACRONYMS = frozenset(
        """msci ucits etf esg sri pab usd eur gbp jpy chf sek nok aud cad em imi ac ii iii
        iv ftse stoxx dax csi jpx nikkei sdg emu us uk eu esm reit reits ai it ihs iboxx
        tips hy ig nr trn dr esgl acwi spdr sp gs cnh rmb kospi asx tsx swap ucit""".split()
    )
    EXPECTED_HEADER = "ShareClass ISIN"
    # The CSV has no asset class column: it has to be inferred (see _infer_class).
    _CASH_MARKERS = re.compile(
        r"\b(cash|deposit|repo|t[- ]?bill|treasury bill|money market|liquidity)\b", re.I
    )
    # DWS uses pseudo-ISINs for non-security positions: "_CURRENCYUSD" for currency,
    # "___ADI2TZ8S1" for index futures.
    _CURRENCY_PREFIX = "_CURRENCY"

    def __init__(self, session: requests.Session):
        self.session = session

    def _download(self, isin: str) -> str | None:
        try:
            response = self.session.get(
                self.EXPORT_URL.format(isin=isin.strip().upper()),
                headers={"Accept": "text/csv,*/*"},
                timeout=TIMEOUT,
            )
        except requests.RequestException as exc:
            log.debug("DWS download failed for %s: %s", isin, exc)
            return None
        # Unknown ISIN -> 500 with an HTML error page.
        if response.status_code != 200:
            return None
        if "csv" not in (response.headers.get("Content-Type") or "").lower():
            return None
        text = response.content.decode("utf-8-sig", errors="replace")
        if not text.lstrip().startswith(self.EXPECTED_HEADER):
            return None
        return text

    def resolve(self, isin: str) -> str | None:
        return self._download(isin)

    def resolve_name(self, isin: str) -> str:
        """Fund name, missing from the CSV export.

        The product page redirects to a slug that contains the name
        (`/en-gb/LU0397221945-portfolio-ucits-etf-1c/`), so reading the final URL is
        enough, instead of downloading and parsing the page.
        """
        try:
            response = self.session.get(
                self.PRODUCT_URL.format(isin=isin.strip().upper()), timeout=TIMEOUT
            )
        except requests.RequestException:
            return ""
        match = re.search(rf"/{re.escape(isin.strip().upper())}-([^/?#]+)", response.url)
        if not match:
            return ""
        words = match.group(1).replace("-", " ").split()
        pretty = " ".join(
            # "1c"/"1d" are share class codes; the rest are acronyms or normal words.
            w.upper() if re.fullmatch(r"\d+[a-z]", w) or w in self._ACRONYMS else w.title()
            for w in words
        )
        return f"Xtrackers {pretty}".strip()

    def fetch(self, isin: str, payload: str | None = None) -> Fund:
        text = payload if payload is not None else self._download(isin)
        if text is None:
            raise ProviderError(f"{isin} does not appear to be an Xtrackers ETF")

        rows = list(csv.DictReader(io.StringIO(text), delimiter=";"))
        default_class = self._fund_default_class(rows)

        holdings = []
        for row in rows:
            name = (row.get("Constituent Name") or "").strip()
            if not name:
                continue
            industry = (row.get("Constituent Industry Classification Name") or "").strip()
            exchange = (row.get("Constituent Main Exchange Name") or "").strip()
            holdings.append(
                Holding(
                    isin=(row.get("Constituent ISIN") or "").strip().upper(),
                    ticker="",  # not exposed by the DWS export
                    name=name,
                    sector_raw=industry,
                    asset_class_raw=self._infer_class(
                        (row.get("Constituent ISIN") or "").strip(),
                        name,
                        industry,
                        default_class,
                    ),
                    country_raw=(row.get("Constituent Country") or "").strip(),
                    currency=(row.get("Constituent Currency ISO Code") or "").strip(),
                    # The DWS export expresses weights as a fraction (0.0558 = 5.58%).
                    weight=_to_float(row.get("Constituent Weighting")) * 100.0,
                    exchange=exchange,
                )
            )

        if not holdings:
            raise ProviderError(f"no constituent returned for {isin}")

        return Fund(isin=isin, name="", issuer=self.name, as_of="", holdings=holdings)

    @staticmethod
    def _fund_default_class(rows: list[dict]) -> str:
        """Prevailing asset class of the fund, inferred from sector coverage.

        DWS populates the sector classification on equity only: on an equity ETF it
        stays "unknown" on a handful of rows, on a fixed income one on almost all of
        them. The share of classified rows therefore tells the two cases apart, and
        drives the leftover rows that do not classify themselves.
        """
        classified = sum(
            1
            for r in rows
            if (r.get("Constituent Industry Classification Name") or "").strip().lower()
            not in ("", "unknown")
        )
        return "Equity" if rows and classified / len(rows) >= 0.5 else "Fixed Income"

    def _infer_class(self, isin: str, name: str, industry: str, default_class: str) -> str:
        """The DWS export has no asset class column: it has to be inferred.

        The exchange is not usable as a signal, because DWS omits it even on listed
        securities (Roche, GSK, AIA). Any reclassification to `Fund` happens upstream,
        when the ISIN resolves as an ETF of a supported issuer.
        """
        if isin.startswith(self._CURRENCY_PREFIX):
            return "Cash"
        if isin and not is_valid_isin(isin):
            return "Derivative"  # pseudo-code: index future
        sector = taxonomy.normalize_sector(industry)
        if sector == taxonomy.CASH_DERIVATIVES or self._CASH_MARKERS.search(name):
            return "Cash"
        # In a fixed income fund the sector, when present, describes the issuer of the
        # security and not the asset class: the fund prevalence takes precedence.
        if default_class != "Equity":
            return default_class
        return "Equity"


# ---------------------------------------------------------------------------
# Vanguard
# ---------------------------------------------------------------------------


class VanguardProvider:
    """Constituents from the GraphQL endpoint behind the Vanguard product pages.

    The product page itself renders only the top ten holdings; the full list comes from
    the same `gpx/graphql` endpoint the page calls, paginated by an opaque cursor.
    """

    name = "Vanguard"

    GRAPHQL_URL = "https://www.vanguard.co.uk/gpx/graphql"
    # Sent by the site on every call; without it the endpoint replies
    # "x-consumer-id must be provided".
    CONSUMER_ID = "uk2"
    PAGE_LIMIT = 1500
    # The largest fund seen is the Global Aggregate Bond ETF at ~13600 holdings, so ten
    # pages; the cap only exists so a broken cursor cannot loop forever.
    MAX_PAGES = 40

    _PROFILE_QUERY = """
        query FundProfile($isins: [String!]) {
          funds(isins: $isins) {
            profile { portId fundFullName fundCurrency }
          }
        }
    """

    # `securityTypes` is deliberately left unset: the site filters it down to equity and
    # bond codes, which drops cash, FX and futures and leaves the weights at ~99.2%.
    # `holdings` and not `delayeredHoldings`: the latter is the issuer's own look-through
    # of a fund of funds, and nested funds are expanded upstream instead.
    # `limit` is inlined rather than passed as a variable: the schema does not declare
    # an `Int` type, so `$limit: Int` fails validation.
    _HOLDINGS_QUERY = """
        query FundsHoldingsQuery($portIds: [String!], $lastItemKey: String) {
          borHoldings(portIds: $portIds) {
            holdings(limit: %d, lastItemKey: $lastItemKey) {
              items {
                issuerName
                securityLongDescription
                isin
                ticker
                securityType
                gicsSectorDescription
                icbIndustryDescription
                marketValuePercentage
                bloombergIsoCountry
                effectiveDate
              }
              lastItemKey
            }
          }
        }
    """ % PAGE_LIMIT

    # Vanguard codes the instrument type rather than the asset class, but the prefix
    # carries it. The values are the raw words `taxonomy.normalize_class` already knows.
    _FUND_TYPES = frozenset({"EQ.ETF", "MF.MF"})
    _CLASS_BY_TYPE = {"CRNY": "Cash"}
    _CLASS_BY_PREFIX = (
        ("EQ.", "Equity"),
        ("FI.", "Fixed Income"),
        ("MM.", "Cash"),        # money market: T-bills, CDs, commercial paper
        ("CT.", "Derivative"),  # FX forwards, spots, portfolio swaps
        ("DE.", "Derivative"),  # index and commodity futures
    )

    def __init__(self, session: requests.Session):
        self.session = session

    def _graphql(self, query: str, variables: dict) -> dict:
        response = self.session.post(
            self.GRAPHQL_URL,
            json={"query": query, "variables": variables},
            headers={
                "Content-Type": "application/json",
                "Accept": "application/json",
                "X-Consumer-ID": self.CONSUMER_ID,
            },
            timeout=TIMEOUT,
        )
        response.raise_for_status()
        payload = response.json()
        if payload.get("errors"):
            message = str(payload["errors"][0].get("message", ""))
            raise ProviderError(f"Vanguard GraphQL error: {message[:200]}")
        return payload.get("data") or {}

    def resolve(self, isin: str) -> dict | None:
        """ISIN -> fund profile, or None when it is not a Vanguard fund.

        One request, and an unknown ISIN comes back as an empty list, so this doubles as
        a cheap issuer probe.
        """
        try:
            data = self._graphql(self._PROFILE_QUERY, {"isins": [isin.strip().upper()]})
        except (requests.RequestException, ValueError, ProviderError) as exc:
            log.debug("Vanguard profile lookup failed for %s: %s", isin, exc)
            return None
        funds = data.get("funds") or []
        profile = (funds[0] or {}).get("profile") if funds else None
        return profile if profile and profile.get("portId") else None

    def fetch(self, isin: str, profile: dict | None = None) -> Fund:
        profile = profile if profile is not None else self.resolve(isin)
        if profile is None:
            raise ProviderError(f"{isin} does not appear to be a Vanguard ETF")
        port_id = str(profile["portId"])

        items: list[dict] = []
        last_item_key = None
        for _ in range(self.MAX_PAGES):
            try:
                data = self._graphql(
                    self._HOLDINGS_QUERY,
                    {"portIds": [port_id], "lastItemKey": last_item_key},
                )
            except (requests.RequestException, ValueError) as exc:
                raise ProviderError(
                    f"Vanguard constituents download failed for {isin}: {exc}"
                ) from exc
            containers = data.get("borHoldings") or []
            page = (containers[0] or {}).get("holdings") if containers else None
            if not page:
                break
            items.extend(page.get("items") or [])
            last_item_key = page.get("lastItemKey")
            if not last_item_key:
                break
        else:
            log.warning("Vanguard: stopped %s after %d pages", isin, self.MAX_PAGES)

        if not items:
            raise ProviderError(f"no constituent returned for {isin}")

        holdings = []
        for item in items:
            holdings.append(
                Holding(
                    isin=(item.get("isin") or "").strip().upper(),
                    ticker=(item.get("ticker") or "").strip(),
                    # Cash and FX rows carry only the long description.
                    name=(item.get("securityLongDescription") or item.get("issuerName") or "").strip(),
                    sector_raw=(
                        item.get("gicsSectorDescription") or item.get("icbIndustryDescription") or ""
                    ).strip(),
                    asset_class_raw=self._infer_class(item.get("securityType")),
                    country_raw=(item.get("bloombergIsoCountry") or "").strip(),
                    currency="",  # no per-holding currency in the schema
                    weight=_to_float(item.get("marketValuePercentage")),
                )
            )

        return Fund(
            isin=isin,
            name=(profile.get("fundFullName") or "").strip(),
            issuer=self.name,
            as_of=(items[0].get("effectiveDate") or "").strip(),
            holdings=holdings,
        )

    @classmethod
    def _infer_class(cls, security_type: str | None) -> str:
        code = (security_type or "").strip().upper()
        if code in cls._FUND_TYPES:
            return "Fund"  # checked first: EQ.ETF would also match the EQ. prefix
        if code in cls._CLASS_BY_TYPE:
            return cls._CLASS_BY_TYPE[code]
        for prefix, asset_class in cls._CLASS_BY_PREFIX:
            if code.startswith(prefix):
                return asset_class
        # Unknown code: pass it through so it surfaces in the unmapped-values warning
        # instead of being silently bucketed.
        return code



# ---------------------------------------------------------------------------
# SPDR / State Street
# ---------------------------------------------------------------------------


class SpdrProvider:
    """Constituents from the daily holdings workbook of the SPDR UCITS range.

    SSGA exposes no per-ISIN lookup: the fund finder behind the product listing returns
    the whole directory in a single response and the site filters it client side. The
    directory is therefore downloaded once and kept, which turns every later probe into
    a dict lookup: the reason this provider goes first in the detection ladder.

    Only the EMEA (UCITS) range is covered. The parallel US feed publishes name, CUSIP
    and SEDOL but no ISIN, country or currency per holding, so its rows could neither be
    aggregated by ISIN nor expanded when they are themselves funds.
    """

    name = "SPDR"

    BASE_URL = "https://www.ssga.com"
    DIRECTORY_URL = BASE_URL + "/bin/v1/ssmp/fund/fundfinder"
    # `de` is the smallest country list that still covers the whole UCITS range: `it`,
    # `fr` and `nl` each miss a white-label sub-fund and `ch` misses a quarter of them.
    # `emea` is a superset, but 8.8 MB and the surplus is the excluded US range.
    DIRECTORY_PARAMS = {
        "country": "de",
        "language": "en_gb",
        "role": "intermediary",
        "product": "etfs",
        "ui": "fund-finder",
    }
    HOLDINGS_DOC_TYPE = "Holdings-daily"
    # The ISIN has no field of its own in the directory: it is one of the search keywords.
    _KEYWORD_ISIN_RE = re.compile(r"\b[A-Z]{2}[A-Z0-9]{9}[0-9]\b")

    # SSGA ships five column layouts (equity, fixed income, a legacy variant, commodity
    # and CLO) that differ in both order and spelling, so columns are located by name.
    # `Identifier` is only an ISIN on the CLO layout; elsewhere it is a CUSIP, and
    # is_valid_isin filters it out.
    _COLUMNS = {
        "isin": ("ISIN", "Identifier"),
        "name": ("Security Name", "Name"),
        "ticker": ("Ticker",),
        "currency": ("Currency", "Currency Local", "Local Currency"),
        "country": ("Trade Country Name", "Country of Issue", "Trade Country"),
        "sector": ("Sector Classification", "Sector"),
        "weight": ("Percent of Fund", "Weight"),
    }
    _WEIGHT_HEADERS = frozenset({"Percent of Fund", "Weight"})
    _HEADER_SCAN_ROWS = 15
    # Only the equity layout carries a sector column; see _infer_class.
    _EQUITY_HEADER = "Sector Classification"
    # The sheet ends with the legal disclaimer in a single cell, a few thousand
    # characters long. No security name comes close.
    _MAX_NAME_LEN = 250

    # The workbook has no asset class column: it has to be inferred (see _infer_class).
    # Both patterns are applied only to rows without a valid ISIN, so a listed security
    # ("Dollar General", "Euronext") cannot be caught by a currency word.
    _DERIVATIVE_MARKERS = re.compile(
        r"^[A-Z]{3}:[A-Z]{3}\b"  # FX forward: "HKD:CNY 20250930"
        r"|\b(?:trs|swap|fut|emini|e[- ]mini|1rty)\b"
        # Index futures carry the delivery month: "EMINI S&P SEP26", "MSCI EAFE SEP6".
        r"|\b(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\d{1,2}\b",
        re.I,
    )
    _CASH_MARKERS = re.compile(
        r"\bcash(?:_|\b)|\b(?:deposit|stif|liquidity|money market)\b"
        # Uninvested FX balances are named after the currency: "Euro", "Swedish Krona".
        r"|\b(?:dollar|euro|pound|sterling|yen|krona|krone|franc|renminbi|yuan|shekel"
        r"|peso|riyal|rial|baht|won|rand|rupiah|ringgit|zloty|lira|dirham|real|forint"
        r"|koruna|rupee|dinar|sol|leu)\b",
        re.I,
    )

    def __init__(self, session: requests.Session):
        self.session = session
        self._entries: dict[str, tuple[str, str]] | None = None

    def _directory(self) -> dict[str, tuple[str, str]]:
        """ISIN -> (fund name, holdings file path), downloaded once per run."""
        if self._entries is None:
            self._entries = self._load_directory()
        return self._entries

    def _load_directory(self) -> dict[str, tuple[str, str]]:
        try:
            response = self.session.get(
                self.DIRECTORY_URL,
                params=self.DIRECTORY_PARAMS,
                headers={"Accept": "application/json"},
                timeout=TIMEOUT,
            )
            response.raise_for_status()
            funds = response.json()["data"]["funds"]["etfs"]["datas"]
        except (requests.RequestException, ValueError, KeyError, TypeError) as exc:
            # A directory that cannot be read degrades into "not a SPDR fund", so the
            # remaining issuers still get their probe.
            log.debug("SSGA fund directory unavailable: %s", exc)
            return {}

        entries: dict[str, tuple[str, str]] = {}
        for item in funds:
            match = self._KEYWORD_ISIN_RE.search(item.get("keywords") or "")
            if not match:
                continue
            path = self._holdings_path(item)
            if path:
                entries.setdefault(match.group(0), (item.get("fundName") or "", path))
        log.debug("SSGA directory: %d funds", len(entries))
        return entries

    @classmethod
    def _holdings_path(cls, item: dict) -> str:
        for group in item.get("documentPdf") or []:
            if group.get("docType") != cls.HOLDINGS_DOC_TYPE:
                continue
            for doc in group.get("docs") or []:
                if doc.get("path"):
                    return str(doc["path"])
        return ""

    def resolve(self, isin: str) -> tuple[str, str] | None:
        return self._directory().get(isin.strip().upper())

    def fetch(self, isin: str, entry: tuple[str, str] | None = None) -> Fund:
        key = isin.strip().upper()
        entry = entry if entry is not None else self.resolve(key)
        if entry is None:
            raise ProviderError(f"{key} does not appear to be a SPDR ETF")
        directory_name, path = entry

        try:
            response = self.session.get(self.BASE_URL + path, timeout=TIMEOUT)
            response.raise_for_status()
        except requests.RequestException as exc:
            raise ProviderError(f"SPDR constituents download failed for {key}: {exc}") from exc

        rows = self._sheet_rows(key, response.content)
        header = self._header_index(rows)
        if header is None:
            raise ProviderError(f"unexpected holdings layout for {key}")

        # The workbook states the ISIN it belongs to, so it validates itself.
        file_isin = self._header_value(rows[:header], "ISIN:").upper()
        if file_isin and file_isin != key:
            raise ProviderError(f"SPDR holdings file for {key} reports {file_isin}")

        columns = [str(c).strip() if c is not None else "" for c in rows[header]]
        at = {
            field_name: next((columns.index(n) for n in names if n in columns), None)
            for field_name, names in self._COLUMNS.items()
        }
        default_class = (
            taxonomy.EQUITY if self._EQUITY_HEADER in columns else taxonomy.FIXED_INCOME
        )

        def cell(row: tuple, field_name: str) -> str:
            index = at[field_name]
            if index is None or index >= len(row) or row[index] is None:
                return ""
            value = str(row[index]).strip()
            # SSGA writes "-" for an empty cell in every column.
            return "" if value == "-" else value

        holdings = []
        for row in rows[header + 1 :]:
            name = cell(row, "name")
            if not name or len(name) > self._MAX_NAME_LEN:
                continue
            constituent = cell(row, "isin").upper()
            if not is_valid_isin(constituent):
                # "-" and "Unassigned" on cash and FX rows, a CUSIP on the commodity one.
                constituent = ""
            holdings.append(
                Holding(
                    isin=constituent,
                    ticker=cell(row, "ticker"),
                    name=name,
                    sector_raw=cell(row, "sector"),
                    asset_class_raw=self._infer_class(constituent, name, default_class),
                    country_raw=cell(row, "country"),
                    currency=cell(row, "currency"),
                    # Already percentage points; "-" on rows rounded down to zero.
                    weight=_to_float(cell(row, "weight")),
                )
            )

        if not holdings:
            raise ProviderError(f"no constituent returned for {key}")

        header_name = self._header_value(rows[:header], "Fund Name:")
        return Fund(
            isin=key,
            name=self._clean_name(header_name or directory_name),
            issuer=self.name,
            as_of=self._header_value(rows[:header], "Holdings As Of:"),
            holdings=holdings,
        )

    @staticmethod
    def _sheet_rows(isin: str, payload: bytes) -> list[tuple]:
        with warnings.catch_warnings():
            # These workbooks carry no default style and openpyxl warns on every load.
            warnings.simplefilter("ignore", UserWarning)
            try:
                book = openpyxl.load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
            except Exception as exc:  # openpyxl surfaces zipfile/XML errors as-is
                raise ProviderError(f"unreadable SPDR workbook for {isin}: {exc}") from exc
            try:
                return list(book.active.iter_rows(values_only=True))
            finally:
                book.close()

    @classmethod
    def _header_index(cls, rows: list[tuple]) -> int | None:
        """Row of the column headers, below the fund name / ISIN / as-of block."""
        for index, row in enumerate(rows[: cls._HEADER_SCAN_ROWS]):
            if {str(c).strip() for c in row if c is not None} & cls._WEIGHT_HEADERS:
                return index
        return None

    @staticmethod
    def _header_value(rows: list[tuple], label: str) -> str:
        for row in rows:
            if row and str(row[0]).strip() == label and len(row) > 1 and row[1] is not None:
                return str(row[1]).strip()
        return ""

    @staticmethod
    def _clean_name(name: str) -> str:
        return re.sub(r"\s+", " ", name.replace("®", "").replace("™", "")).strip()

    @classmethod
    def _infer_class(cls, isin: str, name: str, default_class: str) -> str:
        """The workbook has no asset class column: it has to be inferred.

        The layout gives the prevailing class of the fund (only the equity one carries a
        sector column), and the rows that depart from it are the ones SSGA leaves
        without an ISIN: FX balances, futures and swaps. Any reclassification to `Fund`
        happens upstream, when the ISIN resolves as an ETF of a supported issuer.
        """
        if not isin:
            # Checked first: "EURO STOXX 50 Sep26" would also match the currency words.
            if cls._DERIVATIVE_MARKERS.search(name):
                return taxonomy.DERIVATIVE
            if cls._CASH_MARKERS.search(name):
                return taxonomy.CASH
        return default_class


# ---------------------------------------------------------------------------
# Amundi
# ---------------------------------------------------------------------------


class AmundiProvider:
    """Constituents from the endpoint behind the Amundi ETF product pages.

    The product page renders its holdings table from `composition.compositionData`, which
    `getProductsData` only fills when the request carries a `composition.compositionFields`
    list: without it the field comes back null and the page falls back to the ten-line
    breakdown. The field list below is the one the site itself sends.

    The response ignores the locale for values (sectors and countries are English whatever
    the `languageCode`): the localized spellings in the site's own Excel export are applied
    client side, and do not reach this API.
    """

    name = "Amundi"

    # (host, countryCode, languageCode): a fund is only visible on the sites where it is
    # registered for distribution, and neither list is a superset of the other.
    SITES = (
        ("www.amundietf.fr", "FRA", "fr"),
        ("www.amundietf.co.uk", "GBR", "en"),
    )
    USER_PROFILE = "INSTIT"
    CHARACTERISTICS = ("ISIN", "SHARE_MARKETING_NAME", "POSITION_AS_OF_DATE")
    # `country` is requested because the site requests it, but Amundi never fills it;
    # `countryOfRisk` is the populated one.
    COMPOSITION_FIELDS = (
        "date", "type", "bbg", "isin", "name", "weight",
        "quantity", "currency", "sector", "country", "countryOfRisk",
    )

    # Amundi codes the instrument type rather than the asset class. Money market paper
    # goes to Cash, as it does for Vanguard's `MM.` prefix.
    _CLASS_BY_TYPE = {
        "EQUITY_ORDINARY": "Equity",
        "PREFERENCE_SHARES": "Equity",
        "DEPOSITORY_RECEIPT": "Equity",
        "RIGHT": "Equity",
        "WARRANT": "Equity",
        "CORPORATE": "Fixed Income",
        "GOVERNMENT": "Fixed Income",
        "MEDIUM_TERM_NOTE": "Fixed Income",
        "MUNICIPAL": "Fixed Income",
        "SECURITIZED": "Fixed Income",
        "CASH": "Cash",
        "TREASURY_BILL": "Cash",
        "CERTIFICATE_OF_DEPOSIT": "Cash",
        "FUTURE": "Derivative",
        "STRUCTURED_PRODUCT": "Derivative",
        "ETF": "Fund",
        "MUTUAL_FUND": "Fund",
    }

    def __init__(self, session: requests.Session):
        self.session = session

    def _products(self, host: str, country: str, language: str, isin: str) -> list[dict]:
        payload = {
            "context": {
                "countryCode": country,
                "languageCode": language,
                "userProfileName": self.USER_PROFILE,
            },
            # The productId of an Amundi share class is its ISIN.
            "productIds": [isin],
            "productType": "PRODUCT",
            "characteristics": list(self.CHARACTERISTICS),
            "composition": {"compositionFields": list(self.COMPOSITION_FIELDS)},
        }
        response = self.session.post(
            f"https://{host}/mapi/ProductAPI/getProductsData",
            json=payload,
            headers={"Content-Type": "application/json", "Accept": "application/json"},
            timeout=TIMEOUT,
        )
        response.raise_for_status()
        return response.json().get("products") or []

    def resolve(self, isin: str) -> dict | None:
        """ISIN -> product payload, or None when it is not an Amundi fund.

        One request per site, and an unknown ISIN comes back as an empty list, so this
        doubles as a cheap issuer probe.
        """
        key = isin.strip().upper()
        for host, country, language in self.SITES:
            try:
                products = self._products(host, country, language, key)
            except (requests.RequestException, ValueError) as exc:
                log.debug("Amundi lookup failed for %s on %s: %s", key, host, exc)
                continue
            if products:
                return products[0]
        return None

    def fetch(self, isin: str, product: dict | None = None) -> Fund:
        key = isin.strip().upper()
        product = product if product is not None else self.resolve(key)
        if product is None:
            raise ProviderError(f"{key} does not appear to be an Amundi ETF")

        composition = product.get("composition") or {}
        entries = composition.get("compositionData") or []
        # The whole book comes in one response: `totalNumberOfInstruments` has matched the
        # number of rows on every fund seen, so there is nothing to paginate.
        expected = composition.get("totalNumberOfInstruments")
        if isinstance(expected, int) and expected != len(entries):
            log.warning("Amundi: %s returned %d of %d instruments", key, len(entries), expected)

        holdings = []
        for entry in entries:
            fields = entry.get("compositionCharacteristics") or {}
            security_type = str(fields.get("type") or "").strip()
            currency = str(fields.get("currency") or "").strip()
            name = str(fields.get("name") or "").strip()
            if not name:
                # Amundi leaves the name null on the cash line, and its own Excel export
                # drops the row, quietly losing the weight. Name it after what it is.
                name = " ".join(
                    part for part in (security_type.replace("_", " ").title(), currency) if part
                )
            if not name:
                continue
            # `bbg` is the Bloomberg ticker plus its exchange code ("NVDA UW").
            ticker, _, exchange = str(fields.get("bbg") or "").strip().partition(" ")
            holdings.append(
                Holding(
                    isin=str(fields.get("isin") or "").strip().upper(),
                    ticker=ticker,
                    name=name,
                    sector_raw=str(fields.get("sector") or "").strip(),
                    asset_class_raw=self._infer_class(security_type),
                    country_raw=str(fields.get("countryOfRisk") or "").strip(),
                    currency=currency,
                    # Amundi expresses weights as a fraction (0.0558 = 5.58%).
                    weight=_to_float(entry.get("weight") or fields.get("weight")) * 100.0,
                    exchange=exchange.strip(),
                )
            )

        if not holdings:
            raise ProviderError(f"no constituent returned for {key}")

        characteristics = product.get("characteristics") or {}
        return Fund(
            isin=key,
            name=str(characteristics.get("SHARE_MARKETING_NAME") or "").strip(),
            issuer=self.name,
            as_of=str(characteristics.get("POSITION_AS_OF_DATE") or "").strip(),
            holdings=holdings,
        )

    @classmethod
    def _infer_class(cls, security_type: str | None) -> str:
        code = (security_type or "").strip().upper()
        if code in cls._CLASS_BY_TYPE:
            return cls._CLASS_BY_TYPE[code]
        # Unknown code: pass it through so it surfaces in the unmapped-values warning
        # instead of being silently bucketed.
        return code


# ---------------------------------------------------------------------------
# UBS
# ---------------------------------------------------------------------------


class UbsFileProvider:
    """Constituents from a workbook the user downloaded from the UBS product page.

    UBS is the one issuer that is not fetched. Its holdings sit behind a GraphQL endpoint
    that validates an Azure AD token minted for the product page, and that page is
    geo-restricted: outside the permitted regions it serves an empty placeholder instead
    of the app, so there is nothing to read the token out of. The endpoint carries exactly
    the same six columns as the "Costituenti" download anyway, so reading the file costs
    no information.

    This provider therefore takes a path instead of an ISIN, and stays out of the
    ProviderRegistry probe ladder.
    """

    name = "UBS"

    # The download is OOXML despite its .xls extension; openpyxl refuses the path on the
    # extension alone, so it is always handed a file object.
    _COLUMNS = {
        "name": ("Titolo", "Securities", "Wertpapiere", "Titres"),
        "isin": ("ISIN",),
        "currency": ("Valuta", "Currency", "Währung", "Devise"),
        "weight": ("Ponderazione %", "Weight %", "Gewichtung (%)", "Pondération (%)"),
    }
    # Column order of the export, used when a locale spells a header we do not know.
    _FALLBACK_ORDER = ("name", "isin", None, "currency", None, "weight")
    _ISIN_HEADER = "ISIN"
    _HEADER_SCAN_ROWS = 20
    # ISIN_RE is anchored; the header block embeds the code in a label
    # ("ISIN: : LU0977261329"), so matching inside text needs its own pattern.
    _ISIN_IN_TEXT = re.compile(r"\b[A-Z]{2}[A-Z0-9]{9}[0-9]\b")
    # "Fonte: State Street, 17.08.2026" / "Source: State Street, 17.08.2026"
    _AS_OF_RE = re.compile(r"(\d{2})[.](\d{2})[.](\d{4})")

    def fetch(self, isin: str, path: str) -> Fund:
        key = isin.strip().upper()
        rows = self._sheet_rows(path)

        header = None
        for index, row in enumerate(rows[: self._HEADER_SCAN_ROWS]):
            if any(str(cell).strip() == self._ISIN_HEADER for cell in row if cell):
                header = index
                break
        if header is None:
            raise ProviderError(f"{path}: not a UBS constituents export")

        # The workbook states the fund it belongs to, so it validates itself.
        file_isin = self._find_isin(rows[:header])
        if not file_isin:
            log.warning("%s states no ISIN: cannot confirm it holds %s", path, key)
        elif file_isin != key:
            raise ProviderError(f"{path} holds {file_isin}, not {key}")

        columns = [str(cell).strip() if cell is not None else "" for cell in rows[header]]
        at = {}
        for field_name, names in self._COLUMNS.items():
            index = next((columns.index(n) for n in names if n in columns), None)
            if index is None and field_name in self._FALLBACK_ORDER:
                index = self._FALLBACK_ORDER.index(field_name)
            at[field_name] = index

        def cell(row: tuple, field_name: str) -> str:
            index = at[field_name]
            if index is None or index >= len(row) or row[index] is None:
                return ""
            return str(row[index]).strip()

        data = [row for row in rows[header + 1 :] if is_valid_isin(cell(row, "isin"))]
        if not data:
            raise ProviderError(f"no constituent returned for {key}")

        weights = self._parse_weights([cell(row, "weight") for row in data])

        holdings = []
        for row, weight in zip(data, weights):
            constituent = cell(row, "isin").upper()
            holdings.append(
                Holding(
                    isin=constituent,
                    ticker="",  # the export carries a SEDOL, not a ticker
                    name=cell(row, "name"),
                    sector_raw="",   # UBS publishes neither sector
                    asset_class_raw="",  # nor asset class
                    # No country column either: the ISIN prefix is the country of
                    # registration, which msci resolves directly from the alpha-2 code.
                    country_raw=constituent[:2],
                    currency=cell(row, "currency"),
                    weight=weight,
                )
            )

        return Fund(
            isin=key,
            name=self._fund_name(rows[:header]),
            issuer=self.name,
            as_of=self._as_of(rows[header + 1 :]),
            holdings=holdings,
        )

    @staticmethod
    def _sheet_rows(path: str) -> list[tuple]:
        try:
            payload = open(path, "rb").read()
        except OSError as exc:
            raise ProviderError(f"cannot read {path}: {exc}") from exc
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            try:
                book = openpyxl.load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
            except Exception as exc:  # openpyxl surfaces zipfile/XML errors as-is
                raise ProviderError(f"{path} is not a readable workbook: {exc}") from exc
            try:
                return list(book.active.iter_rows(values_only=True))
            finally:
                book.close()

    @classmethod
    def _parse_weights(cls, raw: list[str]) -> list[float]:
        """Weights, read without having to know the locale of the download.

        The export follows the language of the site it came from: "14,98683" from the
        Italian one, "14.98683" from the English one. Rather than detect that, both
        readings are tried and the one totalling 100 wins - which the column always does.
        """
        european = [cls._number(value, thousands=".", decimal=",") for value in raw]
        anglo = [cls._number(value, thousands=",", decimal=".") for value in raw]
        return min((anglo, european), key=lambda values: abs(sum(values) - 100.0))

    @staticmethod
    def _number(value: str, *, thousands: str, decimal: str) -> float:
        cleaned = value.replace(thousands, "").replace(decimal, ".").replace("%", "").strip()
        try:
            return float(cleaned)
        except ValueError:
            return 0.0

    @classmethod
    def _find_isin(cls, rows: list[tuple]) -> str:
        for row in rows:
            for cell in row:
                match = cls._ISIN_IN_TEXT.search(str(cell).upper()) if cell else None
                if match:
                    return match.group(0)
        return ""

    @classmethod
    def _fund_name(cls, rows: list[tuple]) -> str:
        for row in rows:
            for cell in row:
                text = str(cell).strip() if cell else ""
                # Skip the "ISIN: : LU..." line and anything else that is just a label.
                if text and not cls._ISIN_IN_TEXT.search(text.upper()):
                    return text
        return ""

    @classmethod
    def _as_of(cls, rows: list[tuple]) -> str:
        for row in rows:
            for cell in row:
                match = cls._AS_OF_RE.search(str(cell)) if cell else None
                if match:
                    day, month, year = match.groups()
                    return f"{year}-{month}-{day}"
        return ""


# ---------------------------------------------------------------------------
# Issuer detection
# ---------------------------------------------------------------------------


class ProviderRegistry:
    """Resolve an ISIN to the right provider, caching the attempts.

    The cache also stores the negative outcomes: during recursive expansion the same
    ISIN can recur in several sub-funds and every probe costs an HTTP call.
    """

    def __init__(self, session: requests.Session | None = None):
        self.session = session or build_session()
        self.ishares = IsharesProvider(self.session)
        self.xtrackers = XtrackersProvider(self.session)
        self.vanguard = VanguardProvider(self.session)
        self.spdr = SpdrProvider(self.session)
        self.amundi = AmundiProvider(self.session)
        self._funds: dict[str, Fund] = {}
        self._misses: set[str] = set()

    def fetch(self, isin: str) -> Fund:
        """Download the constituents, cheapest issuer probe first.

        SPDR pays for its whole directory on the first probe and answers from memory
        afterwards, so it costs one request for the entire run and goes first.
        Xtrackers and Vanguard each answer definitively in one request, Amundi in one
        per site; iShares needs a search plus a confirmation per candidate, so it goes last.
        """
        key = isin.strip().upper()
        if key in self._funds:
            return self._funds[key]
        if key in self._misses:
            raise ProviderError(f"{key}: unsupported issuer or non-existent ISIN")

        entry = self.spdr.resolve(key)
        if entry is not None:
            fund = self.spdr.fetch(key, entry=entry)
            self._funds[key] = fund
            return fund

        payload = self.xtrackers.resolve(key)
        if payload is not None:
            fund = self.xtrackers.fetch(key, payload=payload)
            self._funds[key] = fund
            return fund

        profile = self.vanguard.resolve(key)
        if profile is not None:
            fund = self.vanguard.fetch(key, profile=profile)
            self._funds[key] = fund
            return fund

        product = self.amundi.resolve(key)
        if product is not None:
            fund = self.amundi.fetch(key, product=product)
            self._funds[key] = fund
            return fund

        try:
            fund = self.ishares.fetch(key)
        except ProviderError:
            self._misses.add(key)
            raise ProviderError(
                f"{key}: unsupported issuer or non-existent ISIN "
                "(supported: iShares, Xtrackers, Vanguard, SPDR, Amundi)"
            ) from None

        self._funds[key] = fund
        return fund

    def resolve_name(self, fund: Fund) -> str:
        """Fund name, lazily resolved for the providers that do not expose it."""
        if fund.name:
            return fund.name
        if fund.issuer == self.xtrackers.name:
            fund.name = self.xtrackers.resolve_name(fund.isin)
        return fund.name

    def is_supported_etf(self, isin: str) -> bool:
        if not is_valid_isin(isin):
            return False
        try:
            self.fetch(isin)
        except ProviderError:
            return False
        return True


def _to_float(value: str | float | None) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = value.replace(",", "").replace("%", "").strip()
    if not cleaned or cleaned == "-":
        return 0.0
    try:
        return float(cleaned)
    except ValueError:
        return 0.0
